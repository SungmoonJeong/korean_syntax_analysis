"""정답 데이터(문장분석_정답_데이터_0410_수정.xlsx) 파서.

각 문장 블록 구조:
  R    행: A열=번호, B열~ 병합=문장 텍스트
  R+1  행: 형태소 (B열 = 형태소 index 0)
  R+2  행: 영어 글로스
  R+3  행: 품사(영어)
  R+4.. 행: 스팬 레이어. 각 셀/병합셀 = (라벨, 시작열, 끝열)
  마지막: 'Sentence' 라벨(문장 전체)

반환: list of dict {num, text, morphs(list[str]), spans(list[(label,start,end)])}
      start/end 는 형태소 index (0-based, inclusive)
"""
import openpyxl

GOLD_PATH = "docs/문장분석_정답_데이터_0410_수정.xlsx"


def load_gold(path=GOLD_PATH):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["구문분석"]

    # 병합셀: (min_row,min_col) -> max_col  및  좌상단이 아닌 셀 -> 좌상단 매핑
    merged_extent = {}   # (r,c) 좌상단 -> (max_row,max_col)
    for rng in ws.merged_cells.ranges:
        merged_extent[(rng.min_row, rng.min_col)] = (rng.max_row, rng.max_col)

    # 블록 시작행
    starts = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, int) or (isinstance(v, str) and v.strip().isdigit()):
            starts.append((int(v), r))

    results = []
    for bi, (num, R) in enumerate(starts):
        end_r = starts[bi + 1][1] - 1 if bi + 1 < len(starts) else ws.max_row
        text = ws.cell(R, 2).value

        # 형태소 행 = R+1. B열(col2)부터 값이 있는 마지막 열까지.
        morph_row = R + 1
        morphs = []
        last_col = 2
        for c in range(2, ws.max_column + 1):
            v = ws.cell(morph_row, c).value
            if v is not None and str(v).strip() != "":
                last_col = c
        for c in range(2, last_col + 1):
            v = ws.cell(morph_row, c).value
            morphs.append("" if v is None else str(v))

        # 스팬 레이어: R+4 .. end_r
        spans = []
        for r in range(R + 4, end_r + 1):
            c = 2
            while c <= last_col:
                cell = ws.cell(r, c)
                # 병합셀 좌상단인지
                if (r, c) in merged_extent:
                    mr, mc = merged_extent[(r, c)]
                    val = cell.value
                    if val is not None and str(val).strip() != "":
                        label = str(val).strip()
                        spans.append((label, c - 2, min(mc, last_col) - 2))
                    c = mc + 1
                else:
                    val = cell.value
                    if val is not None and str(val).strip() != "":
                        label = str(val).strip()
                        spans.append((label, c - 2, c - 2))
                    c += 1

        results.append({"num": num, "text": text, "morphs": morphs, "spans": spans})
    return results


if __name__ == "__main__":
    import sys, os
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if False else ".")
    data = load_gold(sys.argv[1] if len(sys.argv) > 1 else GOLD_PATH)
    print("문장 수:", len(data))
    for d in data[:2]:
        print("="*70)
        print(f"#{d['num']}: {d['text']}")
        print("형태소:", d["morphs"])
        for lab, s, e in d["spans"]:
            print(f"   [{s:2d}-{e:2d}] {lab:35s} :: {' '.join(d['morphs'][s:e+1])}")
