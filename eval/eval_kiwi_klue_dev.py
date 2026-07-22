"""klue-dev-morph.conllu 기준 Kiwi 태깅 정확도 + UAS/LAS 베이스라인 측정.

목적: core/morph_correction.py(형태소 교정 훅)를 만들기 전에, 아무것도 안 바꾼
지금 파이프라인의 "이전(before)" 수치를 확정하고, 0410_태깅_목록.xlsx의 '문제'
패턴 51건 + kiwi_upgrade/disambiguate의 동형이의어 후보의 실제 빈도·에러율을
dev set(2,000문장, SuPar 학습에 안 쓰인 분할) 기준으로 집계한다.

- 태깅 정확도: 정답 형태소 <-> Kiwi 형태소를 difflib으로 정렬(eval_harness.norm 재사용),
  'equal' 블록으로 정렬된 토큰만 태그 비교(정렬 안 되는 형태소는 별도 집계).
- UAS/LAS: 분절이 정답과 완전히 동일한(1:1) 문장에서만 계산 — 분절이 다르면
  arc 인덱스가 안 맞아 비교 자체가 불가능하기 때문. 해당 문장 비율도 같이 보고.
- 51개 패턴 + 동형이의어 129개 타깃: gold (form,xpos)가 정렬된 위치에서 일치하는지로
  빈도/에러 집계 (rough 1차 우선순위용, 문맥조건까지는 안 봄).

출력: eval/klue_dev_baseline.json
"""
import os
import sys
import re
import json
import difflib
from collections import defaultdict

os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, ".")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from core.analyzer import kiwi_morphs, parse_dep
from eval_harness import load_parser_kiwi, norm

CONLLU_PATH = "supar_morph_dp/klue-dev-morph.conllu"
PROBLEM_XLSX = "../docs/0410_태깅_목록.xlsx"
DISAMBIG_SUMMARY = "../kiwi_upgrade/disambiguate/kiwi_accuracy_summary_final.json"
OUT_PATH = "eval/klue_dev_baseline.json"

PROBLEM_SHEETS = ["보조사, 부사격조사", "종결어미", "연결어미", "선어말어미"]


def load_conllu(path):
    sents = []
    text = None
    rows = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if line.startswith("# text = "):
                text = line[len("# text = "):]
            elif line.startswith("#"):
                continue
            elif line.strip() == "":
                if rows:
                    sents.append({"text": text, "rows": rows})
                rows, text = [], None
            else:
                cols = line.split("\t")
                _id, form, lemma, xpos, upos, _, head, deprel, _, _ = cols
                rows.append({"id": int(_id), "form": form, "xpos": xpos,
                             "head": int(head), "deprel": deprel})
    if rows:
        sents.append({"text": text, "rows": rows})
    return sents


def load_problem_patterns(path):
    patterns = set()
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet in PROBLEM_SHEETS:
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        if "문제" not in header:
            continue
        idx_problem = header.index("문제")
        for row in ws.iter_rows(min_row=2, values_only=True):
            word = row[0]
            problem = row[idx_problem]
            if word and problem and str(problem).strip():
                cleaned = re.sub(r"\(.*?\)", "", str(word)).replace("~", "").strip()
                if cleaned:
                    patterns.add(cleaned)
    return sorted(patterns)


def load_homograph_targets(path):
    with open(path, encoding="utf-8") as f:
        d = json.load(f)
    targets = []
    for item in d["target_breakdown"]:
        info = item["target_info"]
        if "/" not in info:
            continue
        base, tag = info.split("/", 1)
        targets.append((base, tag))
    return targets


def clean_tag(t):
    return t.split("-")[0].split("+")[0] if isinstance(t, str) else t


def main(limit=None):
    print("리소스 로딩 중 (Kiwi + SuPar)...", file=sys.stderr)
    parser, kiwi = load_parser_kiwi()

    sents = load_conllu(CONLLU_PATH)
    if limit:
        sents = sents[:limit]
    print(f"dev 문장 수: {len(sents)}", file=sys.stderr)

    patterns = load_problem_patterns(PROBLEM_XLSX)
    homographs = load_homograph_targets(DISAMBIG_SUMMARY)
    homograph_set = set(homographs)
    print(f"문제 패턴 {len(patterns)}개, 동형이의어 타깃 {len(homographs)}개 로드", file=sys.stderr)

    pattern_freq = defaultdict(int)
    pattern_err = defaultdict(int)
    homo_freq = defaultdict(int)
    homo_err = defaultdict(int)

    total_gold_aligned = 0
    total_gold_correct = 0
    total_unaligned = 0

    uas_num = 0
    las_num = 0
    uas_den = 0
    clean_sent_count = 0

    for si, s in enumerate(sents):
        text = s["text"]
        gold_rows = s["rows"]
        gforms = [r["form"] for r in gold_rows]
        gxpos = [r["xpos"] for r in gold_rows]

        tokens, xpos, spans, lemmas = kiwi_morphs(kiwi, text)

        gn = [norm(m) for m in gforms]
        kn = [norm(m) for m in tokens]
        sm = difflib.SequenceMatcher(a=gn, b=kn, autojunk=False)
        opcodes = sm.get_opcodes()

        for tag, i1, i2, j1, j2 in opcodes:
            if tag == "equal":
                for off in range(i2 - i1):
                    gi, kj = i1 + off, j1 + off
                    total_gold_aligned += 1
                    g_tag = gxpos[gi]
                    k_tag = clean_tag(xpos[kj])
                    correct = (g_tag == k_tag)
                    if correct:
                        total_gold_correct += 1

                    gform = gforms[gi]
                    if gform in patterns:
                        pattern_freq[gform] += 1
                        if not correct:
                            pattern_err[gform] += 1

                    key = (gform, g_tag)
                    if key in homograph_set:
                        homo_freq[key] += 1
                        if not correct:
                            homo_err[key] += 1
            else:
                total_unaligned += (i2 - i1)

        if len(opcodes) == 1 and opcodes[0][0] == "equal" and len(gforms) == len(tokens):
            clean_sent_count += 1
            _, _, arcs, rels = parse_dep(parser, tokens)
            for gi, r in enumerate(gold_rows):
                uas_den += 1
                if arcs[gi] == r["head"]:
                    uas_num += 1
                    if rels[gi] == r["deprel"]:
                        las_num += 1

        if si % 200 == 0:
            print(f"  ...{si}/{len(sents)}", file=sys.stderr)

    result = {
        "n_sentences": len(sents),
        "clean_segmentation_sentences": clean_sent_count,
        "clean_segmentation_rate": (clean_sent_count / len(sents)) if sents else 0,
        "tagging_accuracy": {
            "aligned_tokens": total_gold_aligned,
            "correct": total_gold_correct,
            "accuracy": (total_gold_correct / total_gold_aligned) if total_gold_aligned else 0,
            "unaligned_gold_tokens": total_unaligned,
        },
        "uas": (uas_num / uas_den) if uas_den else None,
        "las": (las_num / uas_den) if uas_den else None,
        "uas_den": uas_den,
        "pattern_stats": sorted(
            [{"pattern": p, "freq": pattern_freq[p], "errors": pattern_err[p]}
             for p in patterns if pattern_freq[p] > 0],
            key=lambda x: -x["errors"],
        ),
        "homograph_stats": sorted(
            [{"base": b, "tag": t, "freq": homo_freq[(b, t)], "errors": homo_err[(b, t)]}
             for (b, t) in homographs if homo_freq[(b, t)] > 0],
            key=lambda x: -x["errors"],
        ),
    }

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(json.dumps({k: v for k, v in result.items()
                       if k not in ("pattern_stats", "homograph_stats")},
                      ensure_ascii=False, indent=2))
    print(f"\n결과 저장: {OUT_PATH}")


if __name__ == "__main__":
    lim = None
    if len(sys.argv) > 1:
        lim = int(sys.argv[1])
    main(limit=lim)
